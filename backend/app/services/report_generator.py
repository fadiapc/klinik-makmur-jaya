import csv
import logging
from datetime import datetime
from pathlib import Path
from typing import Sequence

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.core.config import settings
from app.models.models import Order

logger = logging.getLogger(__name__)

# Klinik Makmur Jaya Brand Colors
PRIMARY_COLOR = (0, 150, 136)  # #009688
PRIMARY_COLOR_HEX = "009688"
TEXT_COLOR = (51, 51, 51)      # #333333

def generate_pdf_report(orders: Sequence[Order], report_dir: Path) -> Path:
    filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    file_path = report_dir / filename

    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.add_page()
    
    # Header / Logo Image
    logo_path = Path("uploads/logo.png")
    if logo_path.exists():
        pdf.image(str(logo_path), x=10, y=10, w=15)
        pdf.set_xy(30, 10)
    else:
        pdf.set_xy(10, 10)
        
    # Header Text
    pdf.set_font("helvetica", "B", 24)
    pdf.set_text_color(*PRIMARY_COLOR)
    pdf.cell(0, 10, "Klinik Makmur Jaya", ln=True, align="L")
    
    # Subtitle
    if logo_path.exists():
        pdf.set_x(30)
    pdf.set_font("helvetica", "", 12)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 6, f"Sales Report - Generated on {datetime.now().strftime('%d %B %Y')}", ln=True, align="L")
    pdf.ln(12)

    # Table Header
    pdf.set_font("helvetica", "B", 10)
    pdf.set_fill_color(*PRIMARY_COLOR)
    pdf.set_text_color(255, 255, 255)
    
    col_widths = [35, 45, 55, 30, 30, 35, 45]
    headers = ["Order Code", "Date", "Customer", "Type", "Status", "Payment", "Total (IDR)"]
    
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], 10, header, border=1, fill=True, align="C")
    pdf.ln()

    # Table Body
    pdf.set_font("helvetica", "", 9)
    pdf.set_text_color(*TEXT_COLOR)
    
    total_revenue = 0.0
    fill = False
    
    for order in orders:
        if fill:
            pdf.set_fill_color(245, 245, 245)
        else:
            pdf.set_fill_color(255, 255, 255)
            
        pdf.cell(col_widths[0], 8, str(order.order_code), border=1, fill=True, align="C")
        pdf.cell(col_widths[1], 8, order.created_at.strftime("%Y-%m-%d %H:%M"), border=1, fill=True, align="C")
        
        customer_name = str(order.customer.full_name) if order.customer else "Unknown"
        # Truncate if too long
        if len(customer_name) > 25:
            customer_name = customer_name[:22] + "..."
            
        pdf.cell(col_widths[2], 8, customer_name, border=1, fill=True, align="L")
        pdf.cell(col_widths[3], 8, str(order.order_type.value), border=1, fill=True, align="C")
        pdf.cell(col_widths[4], 8, str(order.status.value), border=1, fill=True, align="C")
        pdf.cell(col_widths[5], 8, str(order.payment_method.value), border=1, fill=True, align="C")
        
        grand_total = float(order.grand_total)
        pdf.cell(col_widths[6], 8, f"Rp {grand_total:,.0f}".replace(",", "."), border=1, fill=True, align="R")
        pdf.ln()
        
        total_revenue += grand_total
        fill = not fill

    # Footer Total
    pdf.set_font("helvetica", "B", 11)
    pdf.set_fill_color(220, 220, 220)
    pdf.cell(sum(col_widths[:-1]), 10, "GRAND TOTAL REVENUE", border=1, fill=True, align="R")
    pdf.set_text_color(*PRIMARY_COLOR)
    pdf.cell(col_widths[-1], 10, f"Rp {total_revenue:,.0f}".replace(",", "."), border=1, fill=True, align="R")

    pdf.output(str(file_path))
    return file_path


def generate_excel_report(orders: Sequence[Order], report_dir: Path) -> Path:
    filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = report_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "Klinik Makmur Jaya - Sales Report"
    ws["A1"].font = Font(size=18, bold=True, color=PRIMARY_COLOR_HEX)
    
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generated on {datetime.now().strftime('%d %B %Y %H:%M:%S')}"
    ws["A2"].font = Font(size=11, italic=True, color="666666")

    # Table Header
    headers = ["Order Code", "Date", "Customer", "Order Type", "Status", "Payment Method", "Items Count", "Grand Total (IDR)"]
    ws.append([]) # Empty row 3
    ws.append(headers) # Row 4
    
    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(start_color=PRIMARY_COLOR_HEX, end_color=PRIMARY_COLOR_HEX, fill_type="solid")
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    total_revenue = 0.0
    row_idx = 5
    
    for order in orders:
        customer_name = str(order.customer.full_name) if order.customer else "Unknown"
        grand_total = float(order.grand_total)
        
        ws.cell(row=row_idx, column=1, value=str(order.order_code))
        ws.cell(row=row_idx, column=2, value=order.created_at.strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row_idx, column=3, value=customer_name)
        ws.cell(row=row_idx, column=4, value=str(order.order_type.value))
        ws.cell(row=row_idx, column=5, value=str(order.status.value))
        ws.cell(row=row_idx, column=6, value=str(order.payment_method.value))
        ws.cell(row=row_idx, column=7, value=len(order.items))
        ws.cell(row=row_idx, column=8, value=grand_total).number_format = '#,##0'
        
        total_revenue += grand_total
        row_idx += 1

    # Total Row
    ws.merge_cells(f"A{row_idx}:G{row_idx}")
    total_cell_label = ws.cell(row=row_idx, column=1, value="GRAND TOTAL REVENUE")
    total_cell_label.font = Font(bold=True)
    total_cell_label.alignment = Alignment(horizontal="right")
    
    total_val_cell = ws.cell(row=row_idx, column=8, value=total_revenue)
    total_val_cell.font = Font(bold=True, color=PRIMARY_COLOR_HEX)
    total_val_cell.number_format = '#,##0'

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = min(adjusted_width, 40)

    wb.save(file_path)
    return file_path
